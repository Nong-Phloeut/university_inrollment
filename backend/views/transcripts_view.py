from PyQt6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QHBoxLayout, QMessageBox
)
from PyQt6.QtCore import Qt

from controllers.transcript_controller import TranscriptController

from fpdf import FPDF
import openpyxl


class TranscriptsView(QWidget):
    def __init__(self, controller=None):
        super().__init__()
        self.controller = controller
        self.transcript_controller = TranscriptController()
        self.init_ui()
        self.load_transcripts_from_db()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # Title
        title = QLabel("Student Transcripts")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2C3E50;")
        layout.addWidget(title)
        layout.addSpacing(15)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Student Name", "Course Code", "Term", "Grade"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setStyleSheet("""
            QTableWidget {
                font-size: 14px;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                font-weight: bold;
                padding: 5px;
                border: none;
            }
        """)
        layout.addWidget(self.table)

        # Export Buttons
        btn_layout = QHBoxLayout()
        export_pdf_btn = QPushButton("Export to PDF")
        export_pdf_btn.clicked.connect(self.export_pdf)
        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.clicked.connect(self.export_excel)
        btn_layout.addWidget(export_pdf_btn)
        btn_layout.addWidget(export_excel_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.setLayout(layout)
        self.setStyleSheet("background-color: #f9f9f9;")

    def load_transcripts_from_db(self):
        try:
            transcripts = self.transcript_controller.get_all_transcripts()
            self.table.setRowCount(0)
            for transcript in transcripts:
                row = self.table.rowCount()
                self.table.insertRow(row)

                # Extracting fields - adjust keys based on your query result
                student_name = transcript.get('student_name', 'N/A')
                course_code = transcript.get('course_code', 'N/A')
                term = transcript.get('semester', 'N/A')
                grade = transcript.get('grade') or ""  # show blank if None

                self.table.setItem(row, 0, QTableWidgetItem(student_name))
                self.table.setItem(row, 1, QTableWidgetItem(course_code))
                self.table.setItem(row, 2, QTableWidgetItem(term))
                self.table.setItem(row, 3, QTableWidgetItem(grade))
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Could not load transcripts:\n{e}")

    def export_pdf(self):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Student Transcripts", ln=True, align='C')

            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            pdf.cell(200, 10, txt=" | ".join(headers), ln=True)

            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                pdf.cell(200, 10, txt=" | ".join(row_data), ln=True)

            pdf.output("transcripts.pdf")
            QMessageBox.information(self, "Export Complete", "PDF saved as transcripts.pdf")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF:\n{e}")

    def export_excel(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            ws.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save("transcripts.xlsx")
            QMessageBox.information(self, "Export Complete", "Excel file saved as transcripts.xlsx")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel:\n{e}")
